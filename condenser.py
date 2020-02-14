import os
import win32com.client as win32
import openpyxl
import re
import sys

# TODO
# |x| Make tool able to work with just two files
# |x| Make tool compile data onto single worksheet
# | | Modularize methods further
# | | Make CLI for tool
# | | Change tool strings to spanish
# | | Make tool have GUI for files


def main():
    removeXLSXFiles()
    clearLog()
    condense()
    getInventory()


# temporary, to speed up testing
def removeXLSXFiles():
    xlsxFiles = get_files(extension=".xlsx")
    if len(xlsxFiles) > 0:
        confirm = input("Delete the xlsx files in this directory? Y/N: ")
        confirm = confirm.upper()
        if confirm == "Y" or confirm == "YES":
            pass
        else:
            return

    for file in xlsxFiles:
        try:
            os.remove(file)
        except PermissionError:
            print(
                'Couldn\'t erase "%s", please close the document.'
                % file.replace(".\\", "")
            )
            sys.exit()


def clearLog():
    if os.path.exists("log.txt"):
        delFile = input("Clear log? Y/N: ")
        delFile = delFile.upper()
        if delFile == "Y" or delFile == "YES":
            os.remove("log.txt")
            print("Cleared log.")


def writeToLog(line):
    with open("log.txt", "a") as w:
        w.write(line + "\n")


# Reads the generated sheets and then creates a worksheet witht the difference
def getInventory():
    print("Generating Diferences workbook...\n")
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    totVentas = readData("totVentas.xlsx")
    writeToSheet(workbook, totVentas, 0, "Total Ventas")
    print("Creating Total Ventas...\n")
    totCompras = readData("totCompras.xlsx")
    writeToSheet(workbook, totCompras, 1, "Total Compras")
    print("Creating Total Compras...\n")
    inventario = getDifference(totCompras, totVentas)
    writeToSheet(workbook, inventario, 2, "Diferencia Ventas-Compras")
    print("Creating Diferencia Ventas-Compras...\n")
    workbook.save("Diferencias.xlsx")
    print("Done!")


# Writes a dataset to a sheet, formatting the first row and
# setting correct column width.
def writeToSheet(workbook, data, wbPage, wsTitle):
    wb = workbook
    ws = wb.create_sheet(wsTitle, wbPage)
    row = 2
    ws.cell(row=1, column=1, value="Artículo")
    ws.cell(row=1, column=2, value="Descripción")
    ws.cell(row=1, column=3, value="Cantidad")
    for k, v in sorted(data.items()):
        if k != "Artículo":
            ws.cell(row=row, column=1, value=k)
            ws.cell(row=row, column=2, value=v[0])
            ws.cell(row=row, column=3, value=v[1])
        row += 1

    # Set col width
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 10

    # Font and cell color formatting
    greenFill = openpyxl.styles.PatternFill(
        start_color="C0DCC0", end_color="C0DCC0", fill_type="solid"
    )
    ws["A1"].fill = greenFill
    ws["B1"].fill = greenFill
    ws["C1"].fill = greenFill

    bold = openpyxl.styles.Font(bold=True)
    ws["A1"].font = bold
    ws["B1"].font = bold
    ws["C1"].font = bold


def checkDuplicates(*files):
    print("Checking data for duplicate product descriptions...\n")
    fileData = []
    fileDesc = []
    possibleDupes = []
    for f in files:
        for k, v in f.items():
            if k != "Artículo":
                artDeet = delSpacesDetail(v[0])
                fileData.append([k, artDeet])
                fileDesc.append(artDeet)

    # Get items with duplicate descriptions
    n = 0
    for item in fileDesc:
        if fileDesc.count(item) > 1:
            possibleDupes.append(fileData[n])
        n += 1

    # If there are any possible dupes, check them.
    if len(possibleDupes) > 0:
        trueDupes = []
        dupeDict = {}
        trueDupeKeys = []

        for dupe in possibleDupes:
            dupeDesc = dupe[1]
            # Check if possible dupe in dictionary, if not, add it
            if dupeDesc in dupeDict:
                # Compare the keys to see if they match, ignore if they do.
                ddkey = dupeDict[dupeDesc][0]
                if ddkey != dupe[0]:
                    if not dupeDict[dupeDesc][1]:
                        dupeDict[dupeDesc][1] = True
                        duplicate = [ddkey, dupeDesc]
                        trueDupes.append(duplicate)
                        trueDupeKeys.append(ddkey)
                    # Check if key has been accounted for, if not, add.
                    if dupe[0] not in trueDupeKeys:
                        trueDupes.append(dupe)
                        trueDupeKeys.append(dupe[0])
            else:
                dupeDict[dupeDesc] = [dupe[0], False]

        if len(trueDupes) > 0:
            writeToLog("Duplicate descriptions:")
            for n in sorted(trueDupes):
                writeToLog('%s:\t"%s"' % (n[0], n[1]))
            writeToLog("")


# Gets the files and feeds them to the corresponding methods
# Integrate a working directory into the argument list.
def condense():
    if os.path.exists(".\\Diferencias.xlsx"):
        cont = input("\"Diferencias.xlsx\" already exists, continuing will "
                     "delete it, continue? Y/N: ")
        cont = cont.upper()
        if cont == "Y" or cont == "YES":
            os.remove("Diferencias.xlsx")
        else:
            print("Exting...")
            sys.exit()

    print("Transforming xls to xlsx...\n")
    xlsxFileNames = convertAndFormat()
    # Get data from files and condense to totals
    # Compras, Recepciones, Facturas y Remisiones
    print("Reading data...\n")
    count = 0
    compras = {}
    recepciones = {}
    facturas = {}
    remisiones = {}
    for file in xlsxFileNames:
        if "Compras" in file:
            compras = readData(file)
            count += 1
        elif "Recepciones" in file:
            recepciones = readData(file)
            count += 1
        elif "Facturas" in file:
            facturas = readData(file)
            count += 1
        elif "Remisiones" in file:
            remisiones = readData(file)
            count += 1
    # Empty or non-existant files are skipped.
    checkDuplicates(compras, recepciones, facturas, remisiones)
    writeToLog("Done checking for duplicate descriptions, checking duplicate"
               " IDs:")
    writeToLog("")
    # Operations for all 4 files
    if compras and recepciones and facturas and remisiones:
        print("Joining facturas and remisiones...")
        totalCompras = getTotal(facturas, remisiones)
        writeToBook(totalCompras, "Total Ventas", "totVentas.xlsx")

        print("Joining compras and recepciones...")
        totalCompras = getTotal(compras, recepciones)
        writeToBook(totalCompras, "Total Compras", "totCompras.xlsx")

        print("Done generating totCompras.xlsx and totVentas.xlsx\n")

    # xor, make sure we get at least one file of each type
    elif (compras or recepciones) and (facturas or remisiones):
        if facturas and remisiones:
            print("Joining facturas and remisiones...")
        elif facturas:
            print("Using facturas as total ventas")
        else:
            print("Using remisiones as total ventas")

        totalCompras = getTotal(facturas, remisiones)
        writeToBook(totalCompras, "Total Ventas", "totVentas.xlsx")

        if compras and recepciones:
            print("Joining compras and recepciones...")
        elif compras:
            print("Using compras as total compras")
        else:
            print("Using recepciones as total compras")

        totalCompras = getTotal(compras, recepciones)
        writeToBook(totalCompras, "Total Compras", "totCompras.xlsx")

        print("Done generating totCompras.xlsx and totVentas.xlsx\n")
    else:
        print("Not enough files!")
        sys.exit()


def convertAndFormat():
    files = get_files()
    xlsxFileNames = []
    # Convert .xls files to .xlsx and format them
    for filename in files:
        # Check if .xls, if not ignore, if yes, format.
        if filename[-4:] == ".xls":
            xlsxFname = filename + "x"
            if xlsxFname not in files:
                convertToXLSX(filename)
                filename = xlsxFname
                formatFile(xlsxFname)
            else:
                print(
                    "'%s' not converted, there is already a "
                    "file with the same name!" % xlsxFname
                )

            xlsxFileNames.append(xlsxFname)
    return xlsxFileNames


def writeToBook(data, wbTitle, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = wbTitle
    row = 1
    for k, v in data.items():
        ws.cell(row=row, column=1, value=k)
        ws.cell(row=row, column=2, value=v[0])
        ws.cell(row=row, column=3, value=v[1])
        row += 1

    # Set col width
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 10

    wb.save(filename)


# Turn dataB elements negative, then sum to dataA using joinAndTotal
def getDifference(dataA, dataB):
    for key, value in dataB.items():
        dataB[key][1] *= -1
    data = getTotal(dataA, dataB)

    return data


# We use this method so we can iterate over
# the longest dict and make sure we join everything.
def getTotal(dataA, dataB):
    if len(dataA) > len(dataB):
        return joinAndTotal(dataA, dataB)
    else:
        return joinAndTotal(dataB, dataA)


# If len(dataB) == 0 will return A,
# otherwise will join two dicts into A or copy
# everything into A if A is empty.
def joinAndTotal(dataA, dataB):
    printBlank = False
    for key, elem in dataB.items():
        if key in dataA:
            dataA[key][1] = elem[1] + dataA[key][1]
            # Remove dots from name, to avoid false differences
            dataBElem = delSpacesDetail(elem[0])
            dataAElem = delSpacesDetail(dataA[key][0])
            if dataBElem != dataAElem:
                print("Descriptions for %s don't match:" % key)
                print("%s: %s" % (key, dataBElem))
                print("%s: %s" % (key, dataAElem))
                writeToLog("%s: %s" % (key, dataBElem))
                writeToLog("%s: %s" % (key, dataAElem))
                writeToLog("")
                print("")
                printBlank = True
        else:
            dataA[key] = elem
    if printBlank:
        print("")
    return dataA


# Create a dictionary with the data from a formatted file
def readData(filename, keepFile=False):
    if os.path.exists(filename):
        wb = openpyxl.load_workbook(filename)
        ws = wb.active

        dataDict = {}
        for row in ws.iter_rows():
            dataDict[row[0].value] = [row[1].value, row[2].value]
        # Deletes the .xlsx file we created, useful to keep clutter down.
        if not keepFile:
            os.remove(filename)

        return dataDict
    else:
        print("File passed to readData does not exist!")
        return {}


# Deletes unecessary rows and changes row width/height
def formatFile(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    # Delete decorative rows
    ws.delete_rows(1, 2)
    ws.delete_cols(3, 1)
    ws.delete_cols(4, 6)

    row_n = 0
    for row in ws.iter_rows():
        # Set row height
        ws.row_dimensions[row_n].height = 12
        row_n += 1
        # Delete spaces in description
        descrVal = row[1].value
        if descrVal is not None:
            row[1].value = delSpacesDetail(descrVal)
        # Convert numbers to float, delete total rows.
        cant_val = row[2].value
        if row[0].value == "VD":
            ws.delete_rows(row[0].row, 3)
        if row[0].value is None:
            ws.delete_rows(row[0].row, 2)
        elif cant_val != "Cantidad" and cant_val is not None:
            rowStr = row[2].value
            rowStr = rowStr.replace(",", "")
            row[2].value = float(rowStr)
    # Set col width
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 10

    wb.save(filename)


# Removes unecessary spaces from product names
def delSpacesDetail(deetStr):
    deetStr = re.sub(" +", " ", deetStr)
    deetStr = deetStr.replace(". .", "..", -1)
    deetStr = deetStr.replace(" .", ".", -1)
    deetStr = deetStr.replace(".", "", -1)
    return deetStr


# Converts the file from .xls to .xlsx
def convertToXLSX(filename):
    # Replace .\ for current working dir, for windows.
    cwd = os.getcwd()
    filename = filename.replace(".\\", cwd + "\\", 1)
    excel = win32.gencache.EnsureDispatch("Excel.Application")
    wb = excel.Workbooks.Open(filename)

    # FileFormat = 51 is for .xlsx, 56 is for .xls
    wb.SaveAs(filename + "x", FileFormat=51)
    wb.Close()
    excel.Application.Quit()


# Gets the files of a directory
def get_files(path=".", extension=""):
    only_files = []

    for item in os.listdir(path):
        if path[-1] == os.sep:
            path = path[:-1]
        full_path = path + os.sep + item

        if os.path.isfile(full_path):
            if extension:
                f_ext = item[-len(extension):]
                if f_ext == extension:
                    only_files.append(full_path)
            else:
                only_files.append(full_path)

    return only_files


if __name__ == "__main__":
    main()
